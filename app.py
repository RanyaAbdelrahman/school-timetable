import os
import re
import secrets
import tempfile
from datetime import date
from pathlib import Path

import streamlit as st
import pandas as pd
from supabase import create_client, Client

# ============================================================
# إعدادات التطبيق
# ============================================================
st.set_page_config(
    page_title="⭐ نظام الإدارة الذكية للجداول المدرسية ⭐",
    page_icon="🏫",
    layout="centered",
)

# مجلد عمل خاص بكل تشغيل للتطبيق
WORK_DIR = tempfile.mkdtemp(prefix="timetable_")
SCHOOL_NAME = ""

# ============================================================
# Secrets / Supabase
# ============================================================
try:
    SUPABASE_URL = st.secrets["supabase"]["url"]
    SUPABASE_KEY = st.secrets["supabase"]["key"]
except Exception:
    st.error("❌ لم يتم العثور على [supabase] url/key في Streamlit Secrets.")
    st.stop()

ADMIN_PASSWORD = st.secrets.get("ADMIN_PASSWORD")
if not ADMIN_PASSWORD:
    ADMIN_PASSWORD = st.secrets["supabase"].get("ADMIN_PASSWORD")

SERVICE_ROLE_KEY = st.secrets.get("SUPABASE_SERVICE_ROLE_KEY")
if not SERVICE_ROLE_KEY:
    SERVICE_ROLE_KEY = st.secrets["supabase"].get("SUPABASE_SERVICE_ROLE_KEY")

if not ADMIN_PASSWORD:
    st.error("❌ ADMIN_PASSWORD غير موجود في Streamlit Secrets.")
    st.stop()

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
admin_supabase = create_client(SUPABASE_URL, SERVICE_ROLE_KEY) if SERVICE_ROLE_KEY else None

# ============================================================
# أدوات الترخيص
# ============================================================
def normalize_email(email):
    return str(email or "").strip().lower()


def valid_email(email):
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", normalize_email(email)))


def generate_license_key():
    return "CW-" + secrets.token_hex(8).upper()


def rpc_data(result):
    data = result.data
    if isinstance(data, list):
        return data[0] if data else None
    return data


def create_school_request(school_name, email, phone):
    result = supabase.rpc(
        "register_school",
        {
            "p_school_name": school_name.strip(),
            "p_email": normalize_email(email),
            "p_phone": str(phone or "").strip(),
        },
    ).execute()
    data = rpc_data(result)
    if not data:
        return None, False
    return data, not bool(data.get("exists", False))


def get_school(email):
    result = supabase.rpc(
        "check_school_license",
        {"p_email": normalize_email(email)},
    ).execute()
    data = rpc_data(result)
    if not data or not data.get("found", False):
        return None
    return data


def license_state(school):
    if not school:
        return "not_found", "المدرسة غير مسجلة."

    status = str(school.get("status") or "pending").lower().strip()
    if status == "pending":
        return "pending", "طلب التفعيل ما زال قيد المراجعة."
    if status == "rejected":
        return "rejected", "تم رفض طلب تفعيل هذه المدرسة."
    if status in ("blocked", "disabled"):
        return "blocked", "تم إيقاف ترخيص هذه المدرسة."
    if status != "approved":
        return "blocked", "حالة الترخيص غير صالحة."

    start_raw = school.get("start_date")
    expiry_raw = school.get("expiry_date")
    if not start_raw or not expiry_raw:
        return "blocked", "تمت الموافقة ولكن لم يتم تحديد مدة الترخيص بعد."

    try:
        start = date.fromisoformat(str(start_raw)[:10])
        expiry = date.fromisoformat(str(expiry_raw)[:10])
    except Exception:
        return "blocked", "تواريخ الترخيص في قاعدة البيانات غير صحيحة."

    today = date.today()
    if today < start:
        return "not_started", f"الترخيص يبدأ في {start}."
    if today > expiry:
        return "expired", f"انتهت صلاحية الترخيص في {expiry}."
    return "approved", f"الترخيص صالح حتى {expiry}."


def require_admin_client():
    if admin_supabase is None:
        st.error(
            "❌ لوحة الإدارة تحتاج SUPABASE_SERVICE_ROLE_KEY في Streamlit Secrets. "
            "أضفه دون وضعه في GitHub أو داخل app.py."
        )
        st.stop()
    return admin_supabase


def save_license(school_id, start_date, expiry_date, status="approved"):
    client = require_admin_client()
    current = (
        client.table("Schools")
        .select("license_key")
        .eq("id", school_id)
        .limit(1)
        .execute()
    )
    current_data = current.data[0] if current.data else {}
    license_key = current_data.get("license_key") or generate_license_key()
    return (
        client.table("Schools")
        .update({
            "status": status,
            "start_date": str(start_date),
            "expiry_date": str(expiry_date),
            "license_key": license_key,
        })
        .eq("id", school_id)
        .execute()
    ).data

# ============================================================
# CSS
# ============================================================
st.markdown("""
<style>
.stApp { background: linear-gradient(135deg,#f0f4f8 0%,#d9e2ec 100%); font-family:'Cairo','Segoe UI',Tahoma,sans-serif; }
.main-header { background:linear-gradient(135deg,#6366f1 0%,#a855f7 100%); padding:35px; border-radius:20px; color:white; text-align:center; box-shadow:0 10px 25px rgba(99,102,241,.3); margin-bottom:25px; }
.main-header h1 { font-size:30px; font-weight:800; margin-bottom:10px; color:#fff; }
.main-header p { font-size:16px; color:#f3e8ff; margin:0; }
.stTextInput > div > div > input { border-radius:12px; border:2px solid #cbd5e1; padding:12px; font-size:16px; background:#fff; }
.stButton > button { background:linear-gradient(135deg,#10b981 0%,#059669 100%); color:white; font-weight:700; font-size:18px; padding:14px 20px; border-radius:14px; border:none; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# لوحة الإدارة
# ============================================================
with st.sidebar:
    st.markdown("## 🔐 الإدارة")
    admin_mode = st.checkbox("فتح لوحة الإدارة")

if admin_mode:
    st.markdown("# 🔐 لوحة إدارة التراخيص")
    entered_password = st.text_input("كلمة مرور الإدارة", type="password")
    if entered_password != ADMIN_PASSWORD:
        if entered_password:
            st.error("❌ كلمة المرور غير صحيحة.")
        st.info("أدخل كلمة مرور الإدارة لعرض المدارس.")
        st.stop()

    client = require_admin_client()
    st.success("✅ تم تسجيل دخول الإدارة.")
    try:
        schools = client.table("Schools").select("*").order("created_at", desc=True).execute().data or []
    except Exception as e:
        st.error(f"❌ تعذر قراءة جدول Schools: {e}")
        st.stop()

    st.write(f"### 📋 عدد المدارس: {len(schools)}")
    if not schools:
        st.info("لا توجد مدارس مسجلة حتى الآن.")
    for school in schools:
        school_id = school.get("id")
        school_name = school.get("school_name") or "بدون اسم"
        with st.expander(f"🏫 {school_name} — {(school.get('status') or 'pending').upper()}"):
            st.write(f"**Email:** {school.get('email') or ''}")
            st.write(f"**Phone:** {school.get('phone') or ''}")
            st.write(f"**Status:** {school.get('status') or 'pending'}")
            st.write(f"**License:** {school.get('license_key') or '-'}")
            try:
                start_default = date.fromisoformat(str(school.get('start_date'))[:10]) if school.get('start_date') else date.today()
            except Exception:
                start_default = date.today()
            try:
                expiry_default = date.fromisoformat(str(school.get('expiry_date'))[:10]) if school.get('expiry_date') else date.today()
            except Exception:
                expiry_default = date.today()
            c1,c2=st.columns(2)
            with c1: new_start=st.date_input("بداية الترخيص",start_default,key=f"start_{school_id}")
            with c2: new_expiry=st.date_input("نهاية الترخيص",expiry_default,key=f"expiry_{school_id}")
            a,b,c=st.columns(3)
            with a:
                if st.button("✅ اعتماد / تمديد",key=f"approve_{school_id}",use_container_width=True):
                    if new_expiry < new_start: st.error("تاريخ الانتهاء يجب أن يكون بعد تاريخ البداية.")
                    else:
                        try: save_license(school_id,new_start,new_expiry,"approved"); st.success("تم اعتماد/تمديد الترخيص."); st.rerun()
                        except Exception as e: st.error(f"❌ خطأ أثناء الحفظ: {e}")
            with b:
                if st.button("❌ رفض",key=f"reject_{school_id}",use_container_width=True):
                    try: client.table("Schools").update({"status":"rejected"}).eq("id",school_id).execute(); st.rerun()
                    except Exception as e: st.error(f"❌ خطأ: {e}")
            with c:
                if st.button("⛔ إيقاف",key=f"block_{school_id}",use_container_width=True):
                    try: client.table("Schools").update({"status":"blocked"}).eq("id",school_id).execute(); st.rerun()
                    except Exception as e: st.error(f"❌ خطأ: {e}")
    st.stop()

# ============================================================
# بوابة المدرسة
# ============================================================
st.markdown("""
<div class="main-header"><h1> ⭐ نظام الإدارة الذكية للجداول المدرسية ⭐ </h1><p> Code Wonders Academy </p></div>
""", unsafe_allow_html=True)

if "school_verified" not in st.session_state: st.session_state.school_verified=False
if "school_record" not in st.session_state: st.session_state.school_record=None

if not st.session_state.school_verified:
    st.markdown("## 🔐 تفعيل البرنامج")
    school_input_name=st.text_input("🏫 اسم المدرسة",key="school_name_input")
    email_input=st.text_input("📧 البريد الإلكتروني",key="school_email_input")
    phone_input=st.text_input("📱 رقم الهاتف",key="school_phone_input")
    if st.button("📨 التحقق / طلب تفعيل",use_container_width=True):
        if not school_input_name.strip(): st.warning("⚠️ اكتب اسم المدرسة."); st.stop()
        if not valid_email(email_input): st.warning("⚠️ اكتب بريدًا إلكترونيًا صحيحًا."); st.stop()
        try:
            school,created=create_school_request(school_input_name,email_input,phone_input)
            if created:
                st.success("✅ تم إرسال طلب التفعيل بنجاح. انتظر موافقة الإدارة.")
                st.info("يمكنك العودة لاحقًا واستخدام نفس البريد الإلكتروني للتحقق من حالة الطلب.")
                st.stop()
            state,message=license_state(school)
            if state=="approved": st.session_state.school_verified=True; st.session_state.school_record=school; st.rerun()
            elif state=="pending": st.warning("⏳ "+message)
            elif state in ("expired","rejected","blocked"): st.error("❌ "+message)
            else: st.warning("⏳ "+message)
        except Exception as e: st.error(f"❌ حدث خطأ أثناء الاتصال بقاعدة البيانات: {e}")
    st.divider(); st.info("إذا كانت المدرسة مسجلة بالفعل، استخدم نفس البريد الإلكتروني المسجل لدى الإدارة."); st.stop()

try:
    current_school=get_school(st.session_state.school_record.get("email",""))
except Exception as e:
    st.session_state.school_verified=False; st.session_state.school_record=None; st.error(f"❌ تعذر التحقق من الترخيص: {e}"); st.stop()

state,message=license_state(current_school)
if state!="approved":
    st.session_state.school_verified=False; st.session_state.school_record=None
    st.error("❌ "+message if state in ("expired","rejected","blocked") else "⏳ "+message)
    st.stop()

SCHOOL_NAME=current_school.get("school_name") or ""
st.markdown(f"<div style='background:#ecfdf5;padding:12px;border-radius:12px;margin-bottom:15px;'>✅ الترخيص مفعل للمدرسة: <b>{SCHOOL_NAME}</b> | ساري حتى: <b>{current_school.get('expiry_date','')}</b></div>",unsafe_allow_html=True)

if st.button("🚪 تسجيل الخروج من الترخيص"):
    st.session_state.school_verified=False; st.session_state.school_record=None; st.session_state.generated=False; st.rerun()

uploaded_file=st.file_uploader("📂 اختر ملف البيانات بصيغة Excel (inputs.xlsx)",type=["xlsx"])
if uploaded_file is not None:
    if st.button("🚀 إنشاء الجدول المدرسي",use_container_width=True):
        # فحص الترخيص مرة إضافية قبل بدء الحساب
        fresh=get_school(st.session_state.school_record.get("email",""))
        fresh_state,fresh_msg=license_state(fresh)
        if fresh_state!="approved":
            st.error("❌ "+fresh_msg); st.stop()
        input_path=os.path.join(WORK_DIR,"inputs.xlsx")
        with open(input_path,"wb") as f: f.write(uploaded_file.getbuffer())
        with st.spinner("✨ جاري معالجة البيانات وبناء الجداول بدقة، يرجى الانتظار..."):
            try:
                result_path=generate_timetable()
                if result_path and os.path.exists(result_path):
                    st.success("✅ تم توليد الجدول بنجاح!")
                    with open(result_path,"rb") as f: timetable_bytes=f.read()
                    st.download_button("📥 تحميل ملف الجدول النهائي",data=timetable_bytes,file_name="final_timetable.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
                    master_path=get_path("all_classes_master_table.xlsx")
                    if os.path.exists(master_path):
                        with open(master_path,"rb") as f: master_bytes=f.read()
                        st.download_button("📥 تحميل الجدول الشامل لجميع الفصول",data=master_bytes,file_name="all_classes_master_table.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
                else:
                    st.error("❌ لم يتمكن البرنامج من توليد الجدول بسبب وجود قيود متضاربة في ملف الإدخال.")
            except Exception as e:
                st.error(f"❌ حدث خطأ أثناء المعالجة: {e}")


import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from ortools.sat.python import cp_model
import pandas as pd

# ============================================================
# تحديد المسار
# ============================================================


def get_path(filename):
    return os.path.join(WORK_DIR, filename)


SCHOOL_NAME = "مدرسة رواد المستقبل الخاصة بصنبو"


# ============================================================
# تنظيف أيام الإجازة
# ============================================================


def clean_off_days(value):
    if pd.isna(value):
        return []

    text = str(value).strip()

    if not text:
        return []

    # دعم الفاصلة العربية والإنجليزية
    text = text.replace("،", ",")

    result = []

    for day in text.split(","):
        day = day.strip()

        if day and day not in result:
            result.append(day)

    return result


def clean_unwanted_periods(value):
    if pd.isna(value):
        return []
    text = str(value).strip()
    if not text:
        return []
    text = text.replace("،", ",")
    result = []
    for p in text.split(","):
        p = p.strip()
        if p.isdigit():
            val = int(p)
            if val not in result:
                result.append(val)
    return result


# ============================================================
# تنسيق ملف Excel
# ============================================================


def format_excel_workbook(file_path):
    wb = load_workbook(file_path)

    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )

    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

    empty_fill = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )

    data_fill = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )

    off_fill = PatternFill(
        start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
    )

    school_title_font = Font(name="Segoe UI", size=15, bold=True, color="1F4E78")

    section_title_font = Font(name="Segoe UI", size=13, bold=True, color="000000")

    cell_font = Font(name="Segoe UI", size=10, bold=True, color="000000")

    day_font = Font(name="Segoe UI", size=11, bold=True, color="1F4E78")

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_side = Side(style="thin", color="BFBFBF")

    med_side = Side(style="medium", color="1F4E78")

    cell_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]

        ws.views.sheetView[0].showGridLines = True

        # ----------------------------------------------------
        # الشيتات الرئيسية
        # ----------------------------------------------------

        if sheetname in ["Master_Schedule", "كشف_المعلمين", "قائمة_الفصول"]:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            continue

        df_sheet = pd.read_excel(file_path, sheet_name=sheetname)

        num_cols = len(df_sheet.columns)

        if "_" in sheetname:
            sheet_type, title_val = sheetname.split("_", 1)

        else:
            sheet_type = ""
            title_val = sheetname

        if sheet_type == "فصل":
            sub_title = f"جدول حصص فصل: {title_val}"

        elif sheet_type == "مدرس":
            sub_title = f"جدول حصص المعلم/ة: {title_val}"

        elif sheet_type == "قاعة":
            sub_title = f"جدول إشغال قاعة / نشاط: {title_val}"

        else:
            sub_title = title_val

        # ----------------------------------------------------
        # إضافة عنوانين
        # ----------------------------------------------------

        ws.insert_rows(1, amount=2)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)

        ws.cell(row=1, column=1, value=SCHOOL_NAME)

        ws.cell(row=1, column=1).font = school_title_font

        ws.cell(row=1, column=1).alignment = center_align

        ws.row_dimensions[1].height = 25

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)

        ws.cell(row=2, column=1, value=sub_title)

        ws.cell(row=2, column=1).font = section_title_font

        ws.cell(row=2, column=1).alignment = center_align

        ws.row_dimensions[2].height = 22

        # ----------------------------------------------------
        # Header
        # ----------------------------------------------------

        header_row_idx = 3

        ws.row_dimensions[header_row_idx].height = 26

        for col_idx in range(1, num_cols + 1):
            c = ws.cell(row=header_row_idx, column=col_idx)

            c.fill = header_fill
            c.font = header_font
            c.alignment = center_align

            c.border = Border(
                left=thin_side, right=thin_side, top=med_side, bottom=med_side
            )

        # ----------------------------------------------------
        # البيانات
        # ----------------------------------------------------

        max_row = ws.max_row

        for r_idx in range(4, max_row + 1):
            ws.row_dimensions[r_idx].height = 36

            for c_idx in range(1, num_cols + 1):
                cell = ws.cell(row=r_idx, column=c_idx)

                cell.font = cell_font
                cell.alignment = center_align
                cell.border = cell_border
                cell.fill = data_fill

                if c_idx == 1:
                    cell.font = day_font
                    cell.fill = PatternFill(
                        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
                    )
                    
                    # محاذاة تلقائية لعمود اليوم (يمين عربي، يسار إنجليزي)
                    val_text = str(cell.value or "")
                    if any("\u0600" <= char <= "\u06ff" for char in val_text):
                        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                val_str = str(cell.value or "")

                if val_str in ["فراغ", "راحة", "متاحة", "None", ""]:
                    cell.fill = empty_fill

                    cell.font = Font(name="Segoe UI", size=10, italic=True, color="7F7F7F")

                elif "إجازة" in val_str or "OFF" in val_str:
                    cell.fill = off_fill

                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="C00000")

        # ----------------------------------------------------
        # عرض الأعمدة
        # ----------------------------------------------------

        for col in ws.columns:
            max_len = 0

            col_letter = get_column_letter(col[0].column)

            for cell in col:
                val = str(cell.value or "")

                for line in val.split("\n"):
                    if len(line) > max_len:
                        max_len = len(line)

            ws.column_dimensions[col_letter].width = max(max_len + 4, 16)

    wb.save(file_path)


# ============================================================
# توليد الجدول
# ============================================================


def generate_timetable():
    excel_file = get_path("inputs.xlsx")

    if not os.path.exists(excel_file):
        print("❌ لم يتم العثور على ملف inputs.xlsx")
        print(f"المسار: {excel_file}")
        return

    # ========================================================
    # قراءة ملفات Excel
    # ========================================================

    try:
        df_teachers = pd.read_excel(excel_file, sheet_name="Teachers")
        df_classes = pd.read_excel(excel_file, sheet_name="Classes")
        df_assignments = pd.read_excel(excel_file, sheet_name="Assignments")
        df_settings = pd.read_excel(excel_file, sheet_name="Settings")
        df_days = pd.read_excel(excel_file, sheet_name="Days")
    except Exception as e:
        print("❌ حدث خطأ أثناء قراءة inputs.xlsx")
        print(e)
        return

    # ========================================================
    # الأيام
    # ========================================================

    days = [str(d).strip() for d in df_days["DayName"].dropna().tolist()]
    num_days = len(days)

    # ========================================================
    # عدد الحصص
    # ========================================================

    num_periods = int(df_settings["PeriodsPerDay"].iloc[0])

    # ========================================================
    # الفصول
    # ========================================================

    classes = [str(c).strip() for c in df_classes["ClassName"].dropna().tolist()]

    # ========================================================
    # المعلمين
    # ========================================================

    teachers = [str(t).strip() for t in df_teachers["Teacher"].dropna().tolist()]
    periods = [f"الحصة {p + 1}" for p in range(num_periods)]

    # ========================================================
    # OffDays & UnwantedPeriods
    # ========================================================

    teacher_off_days = {}
    teacher_max_off_days = {}
    teacher_unwanted_periods = {}

    for _, row in df_teachers.iterrows():
        if pd.isna(row["Teacher"]):
            continue

        teacher_name = str(row["Teacher"]).strip()

        if "OffDays" in df_teachers.columns:
            off_days = clean_off_days(row["OffDays"])
        else:
            off_days = []

        valid_off_days = []
        for off_day in off_days:
            if off_day in days:
                valid_off_days.append(off_day)
            else:
                print(f"⚠️ اليوم '{off_day}' للمعلم '{teacher_name}' غير موجود في Days")

        teacher_off_days[teacher_name] = valid_off_days

        if "UnwantedPeriods" in df_teachers.columns:
            teacher_unwanted_periods[teacher_name] = clean_unwanted_periods(
                row["UnwantedPeriods"]
            )
        else:
            teacher_unwanted_periods[teacher_name] = []

        if "MaxOffDays" in df_teachers.columns and pd.notna(row.get("MaxOffDays")):
            try:
                teacher_max_off_days[teacher_name] = int(row["MaxOffDays"])
            except (ValueError, TypeError):
                teacher_max_off_days[teacher_name] = len(valid_off_days)
        else:
            teacher_max_off_days[teacher_name] = len(valid_off_days)

    # ========================================================
    # عرض OffDays
    # ========================================================

    print()
    print("=" * 60)
    print("📅 أيام الإجازة لكل معلم")
    print("=" * 60)

    for teacher_name in teachers:
        off_days = teacher_off_days.get(teacher_name, [])
        work_days = [d for d in days if d not in off_days]
        print(f"👤 {teacher_name}")
        print("    OFF: " + (", ".join(off_days) if off_days else "لا يوجد"))
        print(f"    أيام العمل: {len(work_days)} من {num_days}")

    print("=" * 60)

    # ========================================================
    # تنظيف Assignments
    # ========================================================

    clean_assignments = []

    for idx, row in df_assignments.iterrows():
        c = str(row["ClassName"]).strip()
        s = str(row["Subject"]).strip()
        t = str(row["Teacher"]).strip()

        try:
            w = int(row["WeeklyLessons"])
        except (ValueError, TypeError):
            print(f"❌ WeeklyLessons غير صحيح في الصف {idx + 2}")
            return

        r_val = row.get("PreferredRoom", "Classroom")
        if pd.isna(r_val) or str(r_val).strip() in ["nan", "None", ""]:
            r = "Classroom"
        else:
            r = str(r_val).strip()

        clean_assignments.append({
            "idx": idx,
            "c": c,
            "s": s,
            "t": t,
            "r": r,
            "w": w,
        })

    # ========================================================
    # إنشاء Model
    # ========================================================

    model = cp_model.CpModel()
    schedule = {}

    for item in clean_assignments:
        idx = item["idx"]
        c = item["c"]
        s = item["s"]
        t = item["t"]
        r = item["r"]

        for d in range(num_days):
            for p in range(num_periods):
                schedule[(idx, c, s, t, r, d, p)] = model.NewBoolVar(
                    f"var_{idx}_{d}_{p}"
                )

    # ========================================================
    # عدد الحصص الأسبوعية
    # ========================================================

    for item in clean_assignments:
        idx = item["idx"]
        c = item["c"]
        s = item["s"]
        t = item["t"]
        r = item["r"]
        w = item["w"]

        model.Add(
            sum(
                schedule[(idx, c, s, t, r, d, p)]
                for d in range(num_days)
                for p in range(num_periods)
            )
            == w
        )

    # ========================================================
    # منع تعارض الفصل
    # ========================================================

    for c in classes:
        for d in range(num_days):
            for p in range(num_periods):
                relevant_vars = [
                    schedule[(item["idx"], item["c"], item["s"], item["t"], item["r"], d, p)]
                    for item in clean_assignments
                    if c in [x.strip() for x in item["c"].split(",")]
                ]

                if relevant_vars:
                    model.Add(sum(relevant_vars) <= 1)

    # ========================================================
    # منع تعارض المعلم
    # ========================================================

    for teacher_name in teachers:
        for d in range(num_days):
            for p in range(num_periods):
                teacher_vars = []

                for item in clean_assignments:
                    assignment_teachers = [
                        x.strip() for x in item["t"].split("/") if x.strip()
                    ]

                    if teacher_name in assignment_teachers:
                        teacher_vars.append(
                            schedule[(
                                item["idx"],
                                item["c"],
                                item["s"],
                                item["t"],
                                item["r"],
                                d,
                                p,
                            )]
                        )

                if teacher_vars:
                    model.Add(sum(teacher_vars) <= 1)

    # ========================================================
    # OFF DAYS
    # ========================================================

    for item in clean_assignments:
        idx = item["idx"]
        c = item["c"]
        s = item["s"]
        t = item["t"]
        r = item["r"]

        assignment_teachers = [x.strip() for x in t.split("/") if x.strip()]

        for teacher_name in assignment_teachers:
            off_days = teacher_off_days.get(teacher_name, [])

            for d in range(num_days):
                day_name = days[d]

                if day_name in off_days:
                    for p in range(num_periods):
                        model.Add(schedule[(idx, c, s, t, r, d, p)] == 0)

    # ========================================================
    # القاعات
    # ========================================================

    all_rooms = set([
        item["r"]
        for item in clean_assignments
        if item["r"] not in ["Classroom", "nan", ""]
    ])

    for room in all_rooms:
        for d in range(num_days):
            for p in range(num_periods):
                room_vars = [
                    schedule[(item["idx"], item["c"], item["s"], item["t"], item["r"], d, p)]
                    for item in clean_assignments
                    if item["r"] == room
                ]

                if room_vars:
                    model.Add(sum(room_vars) <= 1)

    # ========================================================
    # منع PE / نشاط في الحصة السابعة
    # ========================================================

    target_period_idx = 6

    if num_periods > target_period_idx:
        for item in clean_assignments:
            subject = item["s"]

            if (
                "PE" in subject.upper()
                or "تربيه رياضيه" in subject
                or "تربية رياضية" in subject
            ):
                for d in range(num_days):
                    model.Add(
                        schedule[
                            (
                                item["idx"],
                                item["c"],
                                item["s"],
                                item["t"],
                                item["r"],
                                d,
                                target_period_idx,
                            )
                        ]
                        == 0
                    )

    # ========================================================
    # مزامنة مادة الدين
    # ========================================================

    synced_classes = ["1ب-1", "2ب-1"]
    subject_to_sync = "دين"

    for d in range(num_days):
        for p in range(num_periods):
            var1 = [
                schedule[(item["idx"], item["c"], item["s"], item["t"], item["r"], d, p)]
                for item in clean_assignments
                if (
                    item["c"] == synced_classes[0]
                    and subject_to_sync in item["s"]
                )
            ]

            var2 = [
                schedule[(item["idx"], item["c"], item["s"], item["t"], item["r"], d, p)]
                for item in clean_assignments
                if (
                    item["c"] == synced_classes[1]
                    and subject_to_sync in item["s"]
                )
            ]

            if var1 and var2:
                model.Add(sum(var1) == sum(var2))

    # ========================================================
    # الحد من الحصص المتأخرة
    # ========================================================

    late_periods = [p for p in range(num_periods) if p >= num_periods - 2]

    for item in clean_assignments:
        idx = item["idx"]
        c = item["c"]
        s = item["s"]
        t = item["t"]
        r = item["r"]
        w = item["w"]

        max_allowed_late = max(1, (w + 1) // 2)

        late_vars = [
            schedule[(idx, c, s, t, r, d, p)]
            for d in range(num_days)
            for p in late_periods
        ]

        if late_vars:
            model.Add(sum(late_vars) <= max_allowed_late)

    # ========================================================
    # Objective & Soft Constraints
    # ========================================================

    objective_terms = []
    SUBJECT_DISTRIBUTION_PENALTY = 30
    UNWANTED_PERIOD_PENALTY = 40

    # 1. تفضيل الحصص المبكرة
    for (idx, c, s, t, r, d, p), var in schedule.items():
        weight = (num_periods - p) * 10
        objective_terms.append(var * weight)

    # 2. عدالة توزيع المواد داخل الفصول على الأيام
    class_subject_groups = {}
    for item in clean_assignments:
        class_names = [x.strip() for x in item["c"].split(",") if x.strip()]
        for class_name in class_names:
            key = (class_name, item["s"])
            if key not in class_subject_groups:
                class_subject_groups[key] = []
            class_subject_groups[key].append(item)

    for (class_name, subject), items in class_subject_groups.items():
        daily_load = {}
        for d in range(num_days):
            lesson_vars = []
            for item in items:
                idx = item["idx"]
                c = item["c"]
                s = item["s"]
                t = item["t"]
                r = item["r"]
                item_classes = [x.strip() for x in c.split(",") if x.strip()]

                if class_name in item_classes:
                    for p in range(num_periods):
                        lesson_vars.append(
                            schedule[(idx, c, s, t, r, d, p)]
                        )

            daily_load[d] = model.NewIntVar(
                0, len(lesson_vars), f"subject_load_{class_name}_{subject}_{d}"
            )
            if lesson_vars:
                model.Add(daily_load[d] == sum(lesson_vars))
            else:
                model.Add(daily_load[d] == 0)

        for d1 in range(num_days):
            for d2 in range(d1 + 1, num_days):
                difference = model.NewIntVar(
                    0, num_periods * len(items), f"subject_diff_{class_name}_{subject}_{d1}_{d2}"
                )
                model.AddAbsEquality(difference, daily_load[d1] - daily_load[d2])
                objective_terms.append(difference * (-SUBJECT_DISTRIBUTION_PENALTY))

    # 3. توزيع حصص المواد على الأيام المختلفة
    for item in clean_assignments:
        idx = item["idx"]
        c = item["c"]
        s = item["s"]
        t = item["t"]
        r = item["r"]

        assignment_teachers = [x.strip() for x in t.split("/") if x.strip()]

        for d in range(num_days):
            is_off_day = False
            for teacher_name in assignment_teachers:
                if days[d] in teacher_off_days.get(teacher_name, []):
                    is_off_day = True
                    break

            if not is_off_day:
                day_has_subject = model.NewBoolVar(f"day_has_{idx}_{d}")
                day_lessons = [
                    schedule[(idx, c, s, t, r, d, p)] for p in range(num_periods)
                ]
                model.Add(sum(day_lessons) >= 1).OnlyEnforceIf(day_has_subject)
                model.Add(sum(day_lessons) == 0).OnlyEnforceIf(day_has_subject.Not())
                objective_terms.append(day_has_subject * 80)

    # 4. تفضيل توزيع حصص المعلمين على أيام العمل
    for teacher_name in teachers:
        off_days = teacher_off_days.get(teacher_name, [])
        work_days_indices = [
            d for d in range(num_days) if days[d] not in off_days
        ]

        if not work_days_indices:
            continue

        teacher_day_has_lessons = {}
        for d in work_days_indices:
            teacher_day_has_lessons[d] = model.NewBoolVar(
                f"teacher_active_{teacher_name}_{d}"
            )
            teacher_lessons_on_day = []
            for item in clean_assignments:
                assignment_teachers = [
                    x.strip() for x in item["t"].split("/") if x.strip()
                ]
                if teacher_name in assignment_teachers:
                    for p in range(num_periods):
                        teacher_lessons_on_day.append(
                            schedule[(item["idx"], item["c"], item["s"], item["t"], item["r"], d, p)]
                        )

            if teacher_lessons_on_day:
                model.Add(sum(teacher_lessons_on_day) >= 1).OnlyEnforceIf(
                    teacher_day_has_lessons[d]
                )
                model.Add(sum(teacher_lessons_on_day) == 0).OnlyEnforceIf(
                    teacher_day_has_lessons[d].Not()
                )
                objective_terms.append(teacher_day_has_lessons[d] * 50)

    # 5. تقليل الحصص المتتالية لنفس Assignment
    for item in clean_assignments:
        idx = item["idx"]
        c = item["c"]
        s = item["s"]
        t = item["t"]
        r = item["r"]

        for d in range(num_days):
            for p in range(num_periods - 1):
                both_lessons = model.NewBoolVar(f"both_{idx}_{d}_{p}")
                current_var = schedule[(idx, c, s, t, r, d, p)]
                next_var = schedule[(idx, c, s, t, r, d, p + 1)]

                model.Add(both_lessons <= current_var)
                model.Add(both_lessons <= next_var)
                model.Add(both_lessons >= current_var + next_var - 1)
                objective_terms.append(both_lessons * -5)

    # 6. تفادي الحصص غير المفضلة للمعلمين (ساعات الرضاعة)
    for item in clean_assignments:
        idx = item["idx"]
        c = item["c"]
        s = item["s"]
        t = item["t"]
        r = item["r"]

        assignment_teachers = [x.strip() for x in t.split("/") if x.strip()]

        for teacher_name in assignment_teachers:
            unwanted_ps = teacher_unwanted_periods.get(teacher_name, [])
            if unwanted_ps:
                for d in range(num_days):
                    for p in range(num_periods):
                        period_number = p + 1
                        if period_number in unwanted_ps:
                            objective_terms.append(
                                schedule[(idx, c, s, t, r, d, p)] * (-UNWANTED_PERIOD_PENALTY)
                            )

    # 7. عدالة توزيع حصص المعلمين على أيام العمل (تقليل التباين بين الأيام)
    TEACHER_LOAD_BALANCE_PENALTY = 40  # غرامة التفاوت بين أيام المعلم الواحد

    for teacher_name in teachers:
        off_days = teacher_off_days.get(teacher_name, [])
        work_days_indices = [
            d for d in range(num_days) if days[d] not in off_days
        ]

        if len(work_days_indices) <= 1:
            continue

        teacher_daily_loads = {}
        for d in work_days_indices:
            day_lessons = []
            for item in clean_assignments:
                assignment_teachers = [
                    x.strip() for x in item["t"].split("/") if x.strip()
                ]
                if teacher_name in assignment_teachers:
                    for p in range(num_periods):
                        day_lessons.append(
                            schedule[(item["idx"], item["c"], item["s"], item["t"], item["r"], d, p)]
                        )

            teacher_daily_loads[d] = model.NewIntVar(
                0, num_periods, f"t_load_{teacher_name}_{d}"
            )
            if day_lessons:
                model.Add(teacher_daily_loads[d] == sum(day_lessons))
            else:
                model.Add(teacher_daily_loads[d] == 0)

        for i in range(len(work_days_indices)):
            for j in range(i + 1, len(work_days_indices)):
                d1 = work_days_indices[i]
                d2 = work_days_indices[j]

                diff_var = model.NewIntVar(
                    0, num_periods, f"t_diff_{teacher_name}_{d1}_{d2}"
                )
                model.AddAbsEquality(
                    diff_var, teacher_daily_loads[d1] - teacher_daily_loads[d2]
                )

                objective_terms.append(diff_var * (-TEACHER_LOAD_BALANCE_PENALTY))

    model.Maximize(sum(objective_terms))

    # ========================================================
    # Solver
    # ========================================================

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 60.0
    solver.parameters.num_search_workers = 8

    status = solver.Solve(model)

    # ========================================================
    # توليد ملفات المخرجات
    # ========================================================

    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        print()
        print("=" * 60)
        print("✅ تم توليد الجدول بنجاح!")
        print("=" * 60)

        output_data = []

        for item in clean_assignments:
            idx = item["idx"]
            c = item["c"]
            s = item["s"]
            t = item["t"]
            r = item["r"]

            for d in range(num_days):
                for p in range(num_periods):
                    value = solver.Value(schedule[(idx, c, s, t, r, d, p)])

                    if value == 1:
                        output_data.append({
                            "الفصل": c,
                            "المادة": s,
                            "المدرس": t,
                            "القاعة": r,
                            "اليوم": days[d],
                            "الحصة": f"الحصة {p + 1}",
                        })

        df_result = pd.DataFrame(output_data)
        out_file = get_path("final_timetable.xlsx")

        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            df_result.to_excel(writer, sheet_name="Master_Schedule", index=False)

            summary_rows = []
            for _, row in df_result.iterrows():
                teacher_cell = str(row["المدرس"])
                split_teachers = [t.strip() for t in teacher_cell.split("/") if t.strip()]
                for teacher in split_teachers:
                    summary_rows.append({"المعلم/ة": teacher, "الحصة": 1})

            if summary_rows:
                df_temp = pd.DataFrame(summary_rows)
                df_summary = df_temp.groupby("المعلم/ة")["الحصة"].sum().reset_index()
                df_summary.columns = ["المعلم/ة", "إجمالي الحصص الأسبوعية"]
                df_summary = df_summary.sort_values(by="المعلم/ة").reset_index(drop=True)
            else:
                df_summary = pd.DataFrame(columns=["المعلم/ة", "إجمالي الحصص الأسبوعية"])

            df_summary.to_excel(writer, sheet_name="كشف_المعلمين", index=False)

            df_classes_list = pd.DataFrame({"ClassName": classes})
            df_classes_list.to_excel(writer, sheet_name="قائمة_الفصول", index=False)

            for c in classes:
                df_c = df_result[
                    df_result["الفصل"].apply(
                        lambda x: c in [i.strip() for i in str(x).split(",")]
                    )
                ]

                if not df_c.empty:
                    df_c_copy = df_c.copy()
                    df_c_copy["عرض_الخلايا"] = (
                        df_c_copy["المادة"] + "\n(" + df_c_copy["المدرس"] + ")"
                    )

                    pivot_c = (
                        df_c_copy.pivot_table(
                            index="اليوم",
                            columns="الحصة",
                            values="عرض_الخلايا",
                            aggfunc="first",
                        )
                        .fillna("فراغ")
                        .reindex(index=days, columns=periods)
                    )
                    pivot_c.to_excel(writer, sheet_name=f"فصل_{c}")

            for t in teachers:
                df_t = df_result[
                    df_result["المدرس"].apply(
                        lambda x: t in [i.strip() for i in str(x).split("/")]
                    )
                ]
                off_days_for_t = teacher_off_days.get(t, [])

                if not df_t.empty:
                    df_t_copy = df_t.copy()
                    df_t_copy["عرض_الخلايا"] = (
                        df_t_copy["الفصل"] + "\n(" + df_t_copy["المادة"] + ")"
                    )
                    pivot_t = df_t_copy.pivot_table(
                        index="اليوم", columns="الحصة", values="عرض_الخلايا", aggfunc="first"
                    )
                else:
                    pivot_t = pd.DataFrame(index=days, columns=periods)

                pivot_t = pivot_t.reindex(index=days, columns=periods)
                
                # تعبئة أيام الإجازة والراحة
                for d_idx, day_name in enumerate(days):
                    if day_name in off_days_for_t:
                        for p_col in periods:
                            pivot_t.loc[day_name, p_col] = "إجازة"
                    else:
                        for p_col in periods:
                            if pd.isna(pivot_t.loc[day_name, p_col]):
                                pivot_t.loc[day_name, p_col] = "راحة"

                pivot_t.to_excel(writer, sheet_name=f"مدرس_{t}")

        print("📁 جارٍ تنسيق وتجميل ملف Excel النهائي...")
        format_excel_workbook(out_file)
        print(f"✨ تم الحفظ بنجاح في: {out_file}")
        return out_file

    else:
        print("❌ لم يتم العثور على حل ممكن (Infeasible Model).")
        print("يرجى مراجعة القيود أو الحصص المطلوبة والتأكد من إمكانية جدولتها.")
        return None

