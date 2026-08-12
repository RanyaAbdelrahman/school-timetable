import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def apply_excel_styling(file_path, is_control_panel=False):
    if not os.path.exists(file_path):
        return

    wb = load_workbook(file_path)
    ws = wb.active
    ws.views.sheetView[0].rightToLeft = True

    navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    control_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # أزرق احترافي للتحكم
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # أصفر فاتح مريح للإدخال
    label_fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")  # رمادي مزرق هادئ
    
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    label_font = Font(name="Calibri", size=11, bold=True, color="1F497D")
    input_font = Font(name="Calibri", size=11, bold=True, color="000000")
    normal_font = Font(name="Calibri", size=11)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='B0C4DE'),
        right=Side(style='thin', color='B0C4DE'),
        top=Side(style='thin', color='B0C4DE'),
        bottom=Side(style='thin', color='B0C4DE')
    )

    if is_control_panel:
        # تنسيق ترويسة ملف التحكم
        for col_num in range(1, 3):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = control_header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border
            
        # تنسيق صفوف البيانات داخل ملف التحكم
        for row_num in range(2, ws.max_row + 1):
            cell_label = ws.cell(row=row_num, column=1)
            cell_value = ws.cell(row=row_num, column=2)
            
            cell_label.fill = label_fill
            cell_label.font = label_font
            cell_label.alignment = align_right
            cell_label.border = thin_border
            
            cell_value.fill = input_fill
            cell_value.font = input_font
            cell_value.alignment = align_center
            cell_value.border = thin_border
            
        ws.column_dimensions['A'].width = 42
        ws.column_dimensions['B'].width = 32
        
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 28
    else:
        for cell in ws[1]:
            cell.fill = navy_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border
            
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = normal_font
                cell.alignment = align_center
                cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

    wb.save(file_path)

def generate_substitutions():
    control_filename = "Control_Panel.xlsx"
    timetable_filename = "final_timetable.xlsx"
    
    if not os.path.exists(control_filename):
        df_control = pd.DataFrame({
            'البيان': ['اسم المعلم الغائب', 'اليوم المستهدف', 'القسم أو المواد المطلوبة (اختياري مفصول بـ /)'],
            'القيمة': ['مارينا هشام', 'الخميس', 'عربي / دراسات']
        })
        df_control.to_excel(control_filename, index=False, engine='openpyxl')
        apply_excel_styling(control_filename, is_control_panel=True)

    try:
        df_ctrl = pd.read_excel(control_filename)
        absent_teacher = str(df_ctrl.iloc[0, 1]).strip()
        target_day = str(df_ctrl.iloc[1, 1]).strip()
        
        required_subjects = []
        if len(df_ctrl) > 2:
            val = df_ctrl.iloc[2, 1]
            if pd.notna(val) and str(val).strip().lower() != 'nan':
                raw_subs = str(val).replace(',', '/').split('/')
                required_subjects = [s.strip() for s in raw_subs if s.strip()]
    except Exception as e:
        print(f"[!] خطأ في قراءة ملف التحكم: {e}")
        return

    try:
        df = pd.read_excel(timetable_filename, sheet_name='Master_Schedule')
    except Exception as e:
        print(f"[!] خطأ في قراءة الجدول الرئيسي: {e}")
        return

    df['المدرس'] = df['المدرس'].astype(str).str.strip()
    df['اليوم'] = df['اليوم'].astype(str).str.strip()
    df['المادة'] = df['المادة'].astype(str).str.strip()
    
    absent_classes = df[(df['المدرس'] == absent_teacher) & (df['اليوم'] == target_day)]
    
    if absent_classes.empty:
        print(f"[!] تنبيه: لا توجد حصص مسجلة للمعلم '{absent_teacher}' في يوم '{target_day}'.")
        return

    all_teachers = df['المدرس'].unique()
    substitution_results = []

    for _, row in absent_classes.iterrows():
        period = row['الحصة']
        class_name = row['الفصل']
        subject = row['المادة']

        busy_teachers_in_period = df[(df['اليوم'] == target_day) & (df['الحصة'] == period)]['المدرس'].values
        available_teachers = [t for t in all_teachers if t not in busy_teachers_in_period and t != absent_teacher and pd.notna(t)]

        if required_subjects:
            teachers_with_subjects = df[df['المادة'].isin(required_subjects)]['المدرس'].unique()
            available_teachers = [t for t in available_teachers if t in teachers_with_subjects]

        substitution_results.append({
            'اليوم': target_day,
            'الحصة': period,
            'الفصل الاحتياطي': class_name,
            'المادة': subject,
            'المعلم الغائب': absent_teacher,
            'المعلمون المتاحون (للاحتياطي)': ", ".join(available_teachers)
        })

    result_df = pd.DataFrame(substitution_results)
    output_filename = f"جدول_احتياطي_{absent_teacher}_{target_day}.xlsx"
    result_df.to_excel(output_filename, index=False, engine='openpyxl')
    
    apply_excel_styling(output_filename, is_control_panel=False)
    print(f"[+] تم بنجاح توليد ملف الاحتياطي المنسق: {output_filename}")

if __name__ == "__main__":
    generate_substitutions()