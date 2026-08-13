import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from ortools.sat.python import cp_model
import pandas as pd

# ============================================================
# تحديد المسار
# ============================================================


def get_path(filename):
  if getattr(sys, "frozen", False):
    return os.path.join(os.path.dirname(sys.executable), filename)
  else:
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)


SCHOOL_NAME = "مدرسة رواد المستقبل الخاصة بصنبو"


# ============================================================
# تنظيف أيام الإجازة
# ============================================================


def clean_off_days(value):
  if pd.isna(value):
    return []

  text = str(value).strip()

  if not text:
    return []

  # دعم الفاصلة العربية والإنجليزية
  text = text.replace("،", ",")

  result = []

  for day in text.split(","):
    day = day.strip()

    if day and day not in result:
      result.append(day)

  return result


# ============================================================
# تنسيق ملف Excel
# ============================================================


def format_excel_workbook(file_path):
  wb = load_workbook(file_path)

  header_fill = PatternFill(
      start_color="1F4E78", end_color="1F4E78", fill_type="solid"
  )

  header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

  empty_fill = PatternFill(
      start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
  )

  data_fill = PatternFill(
      start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
  )

  off_fill = PatternFill(
      start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
  )

  school_title_font = Font(name="Segoe UI", size=15, bold=True, color="1F4E78")

  section_title_font = Font(name="Segoe UI", size=13, bold=True, color="000000")

  cell_font = Font(name="Segoe UI", size=10, bold=True, color="000000")

  day_font = Font(name="Segoe UI", size=11, bold=True, color="1F4E78")

  center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

  thin_side = Side(style="thin", color="BFBFBF")

  med_side = Side(style="medium", color="1F4E78")

  cell_border = Border(
      left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
  )

  for sheetname in wb.sheetnames:
    ws = wb[sheetname]

    ws.views.sheetView[0].showGridLines = True

    # ----------------------------------------------------
    # الشيتات الرئيسية
    # ----------------------------------------------------

    if sheetname in ["Master_Schedule", "كشف_المعلمين", "قائمة_الفصول"]:
      for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

      continue

    df_sheet = pd.read_excel(file_path, sheet_name=sheetname)

    num_cols = len(df_sheet.columns)

    if "_" in sheetname:
      sheet_type, title_val = sheetname.split("_", 1)

    else:
      sheet_type = ""
      title_val = sheetname

    if sheet_type == "فصل":
      sub_title = f"جدول حصص فصل: {title_val}"

    elif sheet_type == "مدرس":
      sub_title = f"جدول حصص المعلم/ة: {title_val}"

    elif sheet_type == "قاعة":
      sub_title = f"جدول إشغال قاعة / نشاط: {title_val}"

    else:
      sub_title = title_val

    # ----------------------------------------------------
    # إضافة عنوانين
    # ----------------------------------------------------

    ws.insert_rows(1, amount=2)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)

    ws.cell(row=1, column=1, value=SCHOOL_NAME)

    ws.cell(row=1, column=1).font = school_title_font

    ws.cell(row=1, column=1).alignment = center_align

    ws.row_dimensions[1].height = 25

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)

    ws.cell(row=2, column=1, value=sub_title)

    ws.cell(row=2, column=1).font = section_title_font

    ws.cell(row=2, column=1).alignment = center_align

    ws.row_dimensions[2].height = 22

    # ----------------------------------------------------
    # Header
    # ----------------------------------------------------

    header_row_idx = 3

    ws.row_dimensions[header_row_idx].height = 26

    for col_idx in range(1, num_cols + 1):
      c = ws.cell(row=header_row_idx, column=col_idx)

      c.fill = header_fill
      c.font = header_font
      c.alignment = center_align

      c.border = Border(
          left=thin_side, right=thin_side, top=med_side, bottom=med_side
      )

    # ----------------------------------------------------
    # البيانات
    # ----------------------------------------------------

    max_row = ws.max_row

    for r_idx in range(4, max_row + 1):
      ws.row_dimensions[r_idx].height = 36

      for c_idx in range(1, num_cols + 1):
        cell = ws.cell(row=r_idx, column=c_idx)

        cell.font = cell_font
        cell.alignment = center_align
        cell.border = cell_border
        cell.fill = data_fill

        if c_idx == 1:
          cell.font = day_font

          cell.fill = PatternFill(
              start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
          )

        val_str = str(cell.value or "")

        if val_str in ["فراغ", "راحة", "متاحة", "None", ""]:
          cell.fill = empty_fill

          cell.font = Font(name="Segoe UI", size=10, italic=True, color="7F7F7F")

        elif "إجازة" in val_str or "OFF" in val_str:
          cell.fill = off_fill

          cell.font = Font(name="Segoe UI", size=10, bold=True, color="C00000")

    # ----------------------------------------------------
    # عرض الأعمدة
    # ----------------------------------------------------

    for col in ws.columns:
      max_len = 0

      col_letter = get_column_letter(col[0].column)

      for cell in col:
        val = str(cell.value or "")

        for line in val.split("\n"):
          if len(line) > max_len:
            max_len = len(line)

      ws.column_dimensions[col_letter].width = max(max_len + 4, 16)

  wb.save(file_path)


# ============================================================
# توليد الجدول
# ============================================================


def generate_timetable():
  excel_file = get_path("inputs.xlsx")

  if not os.path.exists(excel_file):
    print("❌ لم يتم العثور على ملف inputs.xlsx")

    print(f"المسار: {excel_file}")

    return

  # ========================================================
  # قراءة ملفات Excel
  # ========================================================

  try:
    df_teachers = pd.read_excel(excel_file, sheet_name="Teachers")

    df_classes = pd.read_excel(excel_file, sheet_name="Classes")

    df_assignments = pd.read_excel(excel_file, sheet_name="Assignments")

    df_settings = pd.read_excel(excel_file, sheet_name="Settings")

    df_days = pd.read_excel(excel_file, sheet_name="Days")

  except Exception as e:
    print("❌ حدث خطأ أثناء قراءة inputs.xlsx")

    print(e)

    return

  # ========================================================
  # الأيام
  # ========================================================

  days = [str(d).strip() for d in df_days["DayName"].dropna().tolist()]

  num_days = len(days)

  # ========================================================
  # عدد الحصص
  # ========================================================

  num_periods = int(df_settings["PeriodsPerDay"].iloc[0])

  # ========================================================
  # الفصول
  # ========================================================

  classes = [str(c).strip() for c in df_classes["ClassName"].dropna().tolist()]

  # ========================================================
  # المعلمين
  # ========================================================

  teachers = [str(t).strip() for t in df_teachers["Teacher"].dropna().tolist()]

  periods = [f"الحصة {p + 1}" for p in range(num_periods)]

  # ========================================================
  # OffDays
  # ========================================================

  teacher_off_days = {}

  teacher_max_off_days = {}

  for _, row in df_teachers.iterrows():
    if pd.isna(row["Teacher"]):
      continue

    teacher_name = str(row["Teacher"]).strip()

    if "OffDays" in df_teachers.columns:
      off_days = clean_off_days(row["OffDays"])

    else:
      off_days = []

    # ----------------------------------------------------
    # التأكد من أن اليوم موجود
    # ----------------------------------------------------

    valid_off_days = []

    for off_day in off_days:
      if off_day in days:
        valid_off_days.append(off_day)

      else:
        print(
            f"⚠️ اليوم '{off_day}' "
            f"للمعلم '{teacher_name}' "
            f"غير موجود في Days"
        )

    teacher_off_days[teacher_name] = valid_off_days

    # ----------------------------------------------------
    # MaxOffDays
    # ----------------------------------------------------

    if "MaxOffDays" in df_teachers.columns and pd.notna(
        row.get("MaxOffDays")
    ):
      try:
        teacher_max_off_days[teacher_name] = int(row["MaxOffDays"])

      except (ValueError, TypeError):
        teacher_max_off_days[teacher_name] = len(valid_off_days)

    else:
      teacher_max_off_days[teacher_name] = len(valid_off_days)

  # ========================================================
  # عرض OffDays
  # ========================================================

  print()
  print("=" * 60)
  print("📅 أيام الإجازة لكل معلم")
  print("=" * 60)

  for teacher_name in teachers:
    off_days = teacher_off_days.get(teacher_name, [])

    work_days = [d for d in days if d not in off_days]

    print(f"👤 {teacher_name}")

    print("   OFF: " + (", ".join(off_days) if off_days else "لا يوجد"))

    print(f"   أيام العمل: {len(work_days)} من {num_days}")

  print("=" * 60)

  # ========================================================
  # تنظيف Assignments
  # ========================================================

  clean_assignments = []

  for idx, row in df_assignments.iterrows():
    c = str(row["ClassName"]).strip()

    s = str(row["Subject"]).strip()

    t = str(row["Teacher"]).strip()

    try:
      w = int(row["WeeklyLessons"])

    except (ValueError, TypeError):
      print(f"❌ WeeklyLessons غير صحيح " f"في الصف {idx + 2}")

      return

    r_val = row.get("PreferredRoom", "Classroom")

    if pd.isna(r_val) or str(r_val).strip() in ["nan", "None", ""]:
      r = "Classroom"

    else:
      r = str(r_val).strip()

    clean_assignments.append({
        "idx": idx,
        "c": c,
        "s": s,
        "t": t,
        "r": r,
        "w": w,
    })

  # ========================================================
  # إنشاء Model
  # ========================================================

  model = cp_model.CpModel()

  schedule = {}

  # ========================================================
  # Variables
  # ========================================================

  for item in clean_assignments:
    idx = item["idx"]
    c = item["c"]
    s = item["s"]
    t = item["t"]
    r = item["r"]

    for d in range(num_days):
      for p in range(num_periods):
        schedule[(idx, c, s, t, r, d, p)] = model.NewBoolVar(
            f"var_{idx}_{d}_{p}"
        )

  # ========================================================
  # عدد الحصص الأسبوعية
  # ========================================================

  for item in clean_assignments:
    idx = item["idx"]
    c = item["c"]
    s = item["s"]
    t = item["t"]
    r = item["r"]
    w = item["w"]

    model.Add(
        sum(
            schedule[(idx, c, s, t, r, d, p)]
            for d in range(num_days)
            for p in range(num_periods)
        )
        == w
    )

  # ========================================================
  # منع تعارض الفصل
  # ========================================================

  for c in classes:
    for d in range(num_days):
      for p in range(num_periods):
        relevant_vars = [
            schedule[(item["idx"], item["c"], item["s"], item["t"], item["r"], d, p)]
            for item in clean_assignments
            if c in [x.strip() for x in item["c"].split(",")]
        ]

        if relevant_vars:
          model.Add(sum(relevant_vars) <= 1)

  # ========================================================
  # ⭐ منع تعارض المعلم
  # ========================================================

  for teacher_name in teachers:
    for d in range(num_days):
      for p in range(num_periods):
        teacher_vars = []

        for item in clean_assignments:
          assignment_teachers = [
              x.strip() for x in item["t"].split("/") if x.strip()
          ]

          if teacher_name in assignment_teachers:
            teacher_vars.append(
                schedule[
                    (
                        item["idx"],
                        item["c"],
                        item["s"],
                        item["t"],
                        item["r"],
                        d,
                        p,
                    )
                ]
            )

        if teacher_vars:
          model.Add(sum(teacher_vars) <= 1)

  # ========================================================
  # ⭐ OFF DAYS
  # ========================================================

  for item in clean_assignments:
    idx = item["idx"]
    c = item["c"]
    s = item["s"]
    t = item["t"]
    r = item["r"]

    assignment_teachers = [x.strip() for x in t.split("/") if x.strip()]

    for teacher_name in assignment_teachers:
      off_days = teacher_off_days.get(teacher_name, [])

      for d in range(num_days):
        day_name = days[d]

        if day_name in off_days:
          for p in range(num_periods):
            model.Add(schedule[(idx, c, s, t, r, d, p)] == 0)

  # ========================================================
  # القاعات
  # ========================================================

  all_rooms = set([
      item["r"]
      for item in clean_assignments
      if item["r"] not in ["Classroom", "nan", ""]
  ])

  for room in all_rooms:
    for d in range(num_days):
      for p in range(num_periods):
        room_vars = [
            schedule[(item["idx"], item["c"], item["s"], item["t"], item["r"], d, p)]
            for item in clean_assignments
            if item["r"] == room
        ]

        if room_vars:
          model.Add(sum(room_vars) <= 1)

  # ========================================================
  # منع PE / نشاط في الحصة السابعة
  # ========================================================

  target_period_idx = 6

  if num_periods > target_period_idx:
    for item in clean_assignments:
      subject = item["s"]

  if (
              "PE" in subject.upper()
              or "تربيه رياضيه" in subject
              or "تربية رياضية" in subject
          ):
        for d in range(num_days):
          model.Add(
              schedule[
                  (
                      item["idx"],
                      item["c"],
                      item["s"],
                      item["t"],
                      item["r"],
                      d,
                      target_period_idx,
                  )
              ]
              == 0
          )

  # ========================================================
  # مزامنة مادة الدين
  # ========================================================

  synced_classes = ["1ب-1", "2ب-1"]

  subject_to_sync = "دين"

  for d in range(num_days):
    for p in range(num_periods):
      var1 = [
          schedule[(item["idx"], item["c"], item["s"], item["t"], item["r"], d, p)]
          for item in clean_assignments
          if (
              item["c"] == synced_classes[0]
              and subject_to_sync in item["s"]
          )
      ]

      var2 = [
          schedule[(item["idx"], item["c"], item["s"], item["t"], item["r"], d, p)]
          for item in clean_assignments
          if (
              item["c"] == synced_classes[1]
              and subject_to_sync in item["s"]
          )
      ]

      if var1 and var2:
        model.Add(sum(var1) == sum(var2))

  # ========================================================
  # الحد من الحصص المتأخرة
  # ========================================================

  late_periods = [p for p in range(num_periods) if p >= num_periods - 2]

  for item in clean_assignments:
    idx = item["idx"]
    c = item["c"]
    s = item["s"]
    t = item["t"]
    r = item["r"]
    w = item["w"]

    max_allowed_late = max(1, (w + 1) // 2)

    late_vars = [
        schedule[(idx, c, s, t, r, d, p)]
        for d in range(num_days)
        for p in late_periods
    ]

    if late_vars:
      model.Add(sum(late_vars) <= max_allowed_late)

  # ========================================================
  # Objective
  # ========================================================

  objective_terms = []

  # --------------------------------------------------------
  # تفضيل الحصص المبكرة
  # --------------------------------------------------------

  for (idx, c, s, t, r, d, p), var in schedule.items():
    weight = (num_periods - p) * 10

    objective_terms.append(var * weight)

  # ========================================================
  # ⭐ تفضيل توزيع كل مادة على أيام مختلفة (التوزيع العادل)
  # ========================================================

  for item in clean_assignments:
    idx = item["idx"]
    c = item["c"]
    s = item["s"]
    t = item["t"]
    r = item["r"]

    assignment_teachers = [x.strip() for x in t.split("/") if x.strip()]

    for d in range(num_days):
      is_off_day = False
      for teacher_name in assignment_teachers:
        if days[d] in teacher_off_days.get(teacher_name, []):
          is_off_day = True
          break

      if not is_off_day:
        day_has_subject = model.NewBoolVar(f"day_has_{idx}_{d}")

        day_lessons = [
            schedule[(idx, c, s, t, r, d, p)] for p in range(num_periods)
        ]

        model.Add(sum(day_lessons) >= 1).OnlyEnforceIf(day_has_subject)

        model.Add(sum(day_lessons) == 0).OnlyEnforceIf(day_has_subject.Not())

        objective_terms.append(day_has_subject * 80)

  # ========================================================
  # ⭐ تفضيل توزيع حصص المعلمين على أيام العمل (عدالة التوزيع)
  # ========================================================

  for teacher_name in teachers:
    off_days = teacher_off_days.get(teacher_name, [])
    work_days_indices = [
        d for d in range(num_days) if days[d] not in off_days
    ]

    if not work_days_indices:
      continue

    teacher_day_has_lessons = {}
    for d in work_days_indices:
      teacher_day_has_lessons[d] = model.NewBoolVar(
          f"teacher_active_{teacher_name}_{d}"
      )

      teacher_lessons_on_day = []
      for item in clean_assignments:
        assignment_teachers = [
            x.strip() for x in item["t"].split("/") if x.strip()
        ]
        if teacher_name in assignment_teachers:
          for p in range(num_periods):
            teacher_lessons_on_day.append(
                schedule[
                    (
                        item["idx"],
                        item["c"],
                        item["s"],
                        item["t"],
                        item["r"],
                        d,
                        p,
                    )
                ]
            )

      if teacher_lessons_on_day:
        model.Add(sum(teacher_lessons_on_day) >= 1).OnlyEnforceIf(
            teacher_day_has_lessons[d]
        )
        model.Add(sum(teacher_lessons_on_day) == 0).OnlyEnforceIf(
            teacher_day_has_lessons[d].Not()
        )

        objective_terms.append(teacher_day_has_lessons[d] * 50)

  # ========================================================
  # محاولة تقليل عدد الحصص المتتالية لنفس Assignment
  # ========================================================

  for item in clean_assignments:
    idx = item["idx"]
    c = item["c"]
    s = item["s"]
    t = item["t"]
    r = item["r"]

    for d in range(num_days):
      for p in range(num_periods - 1):
        both_lessons = model.NewBoolVar(f"both_{idx}_{d}_{p}")

        current_var = schedule[(idx, c, s, t, r, d, p)]

        next_var = schedule[(idx, c, s, t, r, d, p + 1)]

        model.Add(both_lessons <= current_var)

        model.Add(both_lessons <= next_var)

        model.Add(both_lessons >= current_var + next_var - 1)

        objective_terms.append(both_lessons * -5)

  # ========================================================
  # Objective
  # ========================================================

  model.Maximize(sum(objective_terms))

  # ========================================================
  # Solver
  # ========================================================

  solver = cp_model.CpSolver()

  solver.parameters.max_time_in_seconds = 60.0

  solver.parameters.num_search_workers = 8

  status = solver.Solve(model)

  # ========================================================
  # نجاح التوليد
  # ========================================================

  if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
    print()
    print("=" * 60)
    print("✅ تم توليد الجدول بنجاح!")
    print("=" * 60)

    output_data = []

    for item in clean_assignments:
      idx = item["idx"]
      c = item["c"]
      s = item["s"]
      t = item["t"]
      r = item["r"]

      for d in range(num_days):
        for p in range(num_periods):
          value = solver.Value(schedule[(idx, c, s, t, r, d, p)])

          if value == 1:
            output_data.append({
                "الفصل": c,
                "المادة": s,
                "المدرس": t,
                "القاعة": r,
                "اليوم": days[d],
                "الحصة": f"الحصة {p + 1}",
            })

    df_result = pd.DataFrame(output_data)

    out_file = get_path("final_timetable.xlsx")

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
      df_result.to_excel(writer, sheet_name="Master_Schedule", index=False)

      summary_rows = []

      for _, row in df_result.iterrows():
        teacher_cell = str(row["المدرس"])

        split_teachers = [t.strip() for t in teacher_cell.split("/") if t.strip()]

        for teacher in split_teachers:
          summary_rows.append({"المعلم/ة": teacher, "الحصة": 1})

      if summary_rows:
        df_temp = pd.DataFrame(summary_rows)

        df_summary = df_temp.groupby("المعلم/ة")["الحصة"].sum().reset_index()

        df_summary.columns = ["المعلم/ة", "إجمالي الحصص الأسبوعية"]

        df_summary = df_summary.sort_values(by="المعلم/ة").reset_index(
            drop=True
        )

      else:
        df_summary = pd.DataFrame(
            columns=["المعلم/ة", "إجمالي الحصص الأسبوعية"]
        )

      df_summary.to_excel(writer, sheet_name="كشف_المعلمين", index=False)

      for c in classes:
        df_c = df_result[
            df_result["الفصل"].apply(
                lambda x: c in [i.strip() for i in str(x).split(",")]
            )
        ]

        if not df_c.empty:
          df_c_copy = df_c.copy()

          df_c_copy["عرض_الخلايا"] = (
              df_c_copy["المادة"] + "\n(" + df_c_copy["المدرس"] + ")"
          )

          pivot_c = (
              df_c_copy.pivot_table(
                  index="اليوم",
                  columns="الحصة",
                  values="عرض_الخلايا",
                  aggfunc="first",
              )
              .fillna("فراغ")
              .reindex(index=days, columns=periods)
          )

          pivot_c.to_excel(writer, sheet_name=f"فصل_{c}")

      for t in teachers:
        df_t = df_result[
            df_result["المدرس"].apply(
                lambda x: t in [i.strip() for i in str(x).split("/")]
            )
        ]

        off_days_for_t = teacher_off_days.get(t, [])

        if not df_t.empty:
          df_t_copy = df_t.copy()

          df_t_copy["عرض_الخلايا"] = (
              df_t_copy["الفصل"] + "\n(" + df_t_copy["المادة"] + ")"
          )

          pivot_t = df_t_copy.pivot_table(
              index="اليوم", columns="الحصة", values="عرض_الخلايا", aggfunc="first"
          )

        else:
          pivot_t = pd.DataFrame("راحة", index=days, columns=periods)

        pivot_t = pivot_t.reindex(index=days, columns=periods).fillna("راحة")

        for off_day in off_days_for_t:
          if off_day in pivot_t.index:
            pivot_t.loc[off_day, :] = "إجازة (OFF)"

        pivot_t.to_excel(writer, sheet_name=f"مدرس_{t}")

      for room in sorted(list(all_rooms)):
        df_r = df_result[df_result["القاعة"] == room]

        if not df_r.empty:
          df_r_copy = df_r.copy()

          df_r_copy["عرض_الخلايا"] = (
              df_r_copy["الفصل"] + "\n(" + df_r_copy["المادة"] + ")"
          )

          pivot_r = (
              df_r_copy.pivot_table(
                  index="اليوم",
                  columns="الحصة",
                  values="عرض_الخلايا",
                  aggfunc="first",
              )
              .fillna("متاحة")
              .reindex(index=days, columns=periods)
          )

          pivot_r.to_excel(writer, sheet_name=f"قاعة_{room}")
# ========================================================
# ========================================================
    # ⭐ توليد ملف مستقل لشامل جداول جميع الفصول في شيت واحد (بتصميم احترافي وألوان جذابة)
    # ========================================================
    all_classes_matrix_file = "all_classes_master_table.xlsx"

    with pd.ExcelWriter(all_classes_matrix_file, engine="openpyxl") as writer:
      wb = writer.book
      ws = wb.create_sheet(title="جداول_جميع_الفصول")
      if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

      ws.views.sheetView[0].showGridLines = True
      ws.sheet_view.rightToLeft = True

      # الألوان والتنسيقات المحدثة (درجات الأزرق والاحترافي)
      thin_border = Border(
          left=Side(style="thin", color="D3D3D3"),
          right=Side(style="thin", color="D3D3D3"),
          top=Side(style="thin", color="D3D3D3"),
          bottom=Side(style="thin", color="D3D3D3"),
      )
      header_fill = PatternFill(
          start_color="1F4E78", end_color="1F4E78", fill_type="solid"
      )  # أزرق غامق جذاب
      sub_header_fill = PatternFill(
          start_color="2F5597", end_color="2F5597", fill_type="solid"
      )  # أزرق متوسط
      title_fill = PatternFill(
          start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
      )  # أزرق فاتح للعنوان

      center_align = Alignment(
          horizontal="center", vertical="center", wrap_text=True
      )
      title_font = Font(name="Segoe UI", size=14, bold=True, color="1F4E78")
      header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")

      # 1. إضافة عنوان المدرسة الرئيسي في السطر الأول
      ws.merge_cells("A1:V1")
      title_cell = ws.cell(
          row=1, column=1, value="جدول الحصص المدرسي المجمع - لجميع الفصول"
      )
      title_cell.font = title_font
      title_cell.alignment = center_align
      title_cell.fill = title_fill
      ws.row_dimensions[1].height = 40

      # ترك سطر فارغ (السطر رقم 2) للتنسيق البصري الجميل

      # 2. بناء الهيدر الرئيسي (يبدأ من السطر الثالث 3 و 4)
      header_row1 = 3
      header_row2 = 4

      ws.row_dimensions[header_row1].height = 25
      ws.row_dimensions[header_row2].height = 22

      # عمود المسلسل
      ws.cell(row=header_row1, column=1, value="م").font = header_font
      ws.cell(row=header_row1, column=1).alignment = center_align
      ws.cell(row=header_row1, column=1).fill = header_fill
      ws.cell(row=header_row1, column=1).border = thin_border
      ws.merge_cells(
          start_row=header_row1,
          start_column=1,
          end_row=header_row2,
          end_column=1,
      )

      # عمود اسم الفصل
      ws.cell(row=header_row1, column=2, value="اسم الفصل").font = header_font
      ws.cell(row=header_row1, column=2).alignment = center_align
      ws.cell(row=header_row1, column=2).fill = header_fill
      ws.cell(row=header_row1, column=2).border = thin_border
      ws.merge_cells(
          start_row=header_row1,
          start_column=2,
          end_row=header_row2,
          end_column=2,
      )

      current_col = 3
      for day in days:
        start_c = current_col
        end_c = current_col + len(periods) - 1

        # دمج اسم اليوم
        ws.merge_cells(
            start_row=header_row1,
            start_column=start_c,
            end_row=header_row1,
            end_column=end_c,
        )
        day_cell = ws.cell(row=header_row1, column=start_c, value=day)
        day_cell.font = header_font
        day_cell.alignment = center_align
        day_cell.fill = header_fill

        # تخطيط حدود خلايا اليوم المدمجة
        for c in range(start_c, end_c + 1):
          ws.cell(row=header_row1, column=c).border = thin_border

        # أرقام الحصص تحت كل يوم
        for p_idx, p in enumerate(periods):
          p_cell = ws.cell(
              row=header_row2, column=start_c + p_idx, value=f"ح {p}"
          )
          p_cell.font = header_font
          p_cell.alignment = center_align
          p_cell.fill = sub_header_fill
          p_cell.border = thin_border

        current_col += len(periods)

      # 3. تعبئة البيانات لكل فصل بتبديل ألوان الصفوف (Zebra Striping)
      row_start_data = 5
      for idx, cls in enumerate(sorted(list(classes)), start=1):
        row_num = row_start_data + idx - 1
        ws.row_dimensions[row_num].height = 35

        # تلوين متناوب للصفوف لتسهيل القراءة (أبيض ورصاصي فاتح جداً)
        row_fill = (
            PatternFill(
                start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
            )
            if idx % 2 == 0
            else PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            )
        )

        # رقم المسلسل
        c1 = ws.cell(row=row_num, column=1, value=idx)
        c1.alignment = center_align
        c1.border = thin_border
        c1.fill = row_fill
        c1.font = Font(name="Segoe UI", size=9)

        # اسم الفصل
        c2 = ws.cell(row=row_num, column=2, value=str(cls))
        c2.alignment = center_align
        c2.border = thin_border
        c2.fill = row_fill
        c2.font = Font(name="Segoe UI", size=10, bold=True, color="1F4E78")

        # تعبئة الحصص
        col_cursor = 3
        for day in days:
          for p in periods:
            cell_val = "-"
            if not df_r.empty:  # أو استخدام df_result حسب المتغير المتاح لديك
              pass

            # جلب مادة الفصل في هذا اليوم وهذه الحصة
            match = df_result[
                (df_result["الفصل"] == cls)
                & (df_result["اليوم"] == day)
                & (df_result["الحصة"] == p)
            ]
            if not match.empty:
              mat = match.iloc[0]["المادة"]
              tch = match.iloc[0].get("المعلم", "")
              cell_val = f"{mat}\n({tch})" if tch else mat
            else:
              cell_val = "متاحة"

            cell = ws.cell(row=row_num, column=col_cursor, value=cell_val)
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = row_fill
            cell.font = Font(name="Segoe UI", size=8)
            col_cursor += 1

    print(f"📁 تم إنشاء ملف جداول الفصول المجمع بتنسيق جذاب: {all_classes_matrix_file}")
    format_excel_workbook(out_file)

    print("💾 تم حفظ الملف بنجاح:")
    print(out_file)
  else:
    print()
    print("=" * 60)
    print("❌ لم يتمكن البرنامج من توليد الجدول")
    print("=" * 60)
# ========================================================
    # ⭐ 1. تجميع المواد لكل معلم وحساب نطاق الحصص للتلوين
    # ========================================================
    teacher_subjects = {}
    for t in teachers:
      subs = (
          df_result[df_result["المعلم"] == t]["المادة"].dropna().unique().tolist()
      )
      teacher_subjects[t] = ", ".join(subs)

    counts = [len(df_result[df_result["المعلم"] == t]) for t in teachers]
    max_c = max(counts) if counts else 1
    min_c = min(counts) if counts else 1
    range_c = max_c - min_c if max_c != min_c else 1

    # ========================================================
    # ⭐ 2. توليد ملف مستقل لشامل جداول جميع الفصول في شيت واحد بتنسيق جذاب
    # ========================================================
    all_classes_matrix_file = "all_classes_master_table.xlsx"

    with pd.ExcelWriter(all_classes_matrix_file, engine="openpyxl") as writer:
      wb = writer.book
      ws = wb.create_sheet(title="جداول_جميع_الفصول")
      if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

      ws.views.sheetView[0].showGridLines = True
      ws.sheet_view.rightToLeft = True

      thin_border = Border(
          left=Side(style="thin", color="D3D3D3"),
          right=Side(style="thin", color="D3D3D3"),
          top=Side(style="thin", color="D3D3D3"),
          bottom=Side(style="thin", color="D3D3D3"),
      )
      header_fill = PatternFill(
          start_color="1F4E78", end_color="1F4E78", fill_type="solid"
      )
      sub_header_fill = PatternFill(
          start_color="2F5597", end_color="2F5597", fill_type="solid"
      )
      title_fill = PatternFill(
          start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
      )

      center_align = Alignment(
          horizontal="center", vertical="center", wrap_text=True
      )
      title_font = Font(name="Segoe UI", size=14, bold=True, color="1F4E78")
      header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")

      # عنوان المدرسة الرئيسي
      ws.merge_cells("A1:V1")
      title_cell = ws.cell(
          row=1, column=1, value="جدول الحصص المدرسي المجمع - لجميع الفصول"
      )
      title_cell.font = title_font
      title_cell.alignment = center_align
      title_cell.fill = title_fill
      ws.row_dimensions[1].height = 40

      header_row1 = 3
      header_row2 = 4

      ws.row_dimensions[header_row1].height = 25
      ws.row_dimensions[header_row2].height = 22

      ws.cell(row=header_row1, column=1, value="م").font = header_font
      ws.cell(row=header_row1, column=1).alignment = center_align
      ws.cell(row=header_row1, column=1).fill = header_fill
      ws.cell(row=header_row1, column=1).border = thin_border
      ws.merge_cells(
          start_row=header_row1,
          start_column=1,
          end_row=header_row2,
          end_column=1,
      )

      ws.cell(row=header_row1, column=2, value="اسم الفصل").font = header_font
      ws.cell(row=header_row1, column=2).alignment = center_align
      ws.cell(row=header_row1, column=2).fill = header_fill
      ws.cell(row=header_row1, column=2).border = thin_border
      ws.merge_cells(
          start_row=header_row1,
          start_column=2,
          end_row=header_row2,
          end_column=2,
      )

      current_col = 3
      for day in days:
        start_c = current_col
        end_c = current_col + len(periods) - 1

        ws.merge_cells(
            start_row=header_row1,
            start_column=start_c,
            end_row=header_row1,
            end_column=end_c,
        )
        day_cell = ws.cell(row=header_row1, column=start_c, value=day)
        day_cell.font = header_font
        day_cell.alignment = center_align
        day_cell.fill = header_fill

        for c in range(start_c, end_c + 1):
          ws.cell(row=header_row1, column=c).border = thin_border

        for p_idx, p in enumerate(periods):
          p_cell = ws.cell(
              row=header_row2, column=start_c + p_idx, value=f"ح {p}"
          )
          p_cell.font = header_font
          p_cell.alignment = center_align
          p_cell.fill = sub_header_fill
          p_cell.border = thin_border

        current_col += len(periods)

      row_start_data = 5
      for idx, cls in enumerate(sorted(list(classes)), start=1):
        row_num = row_start_data + idx - 1
        ws.row_dimensions[row_num].height = 35

        row_fill = (
            PatternFill(
                start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
            )
            if idx % 2 == 0
            else PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            )
        )

        c1 = ws.cell(row=row_num, column=1, value=idx)
        c1.alignment = center_align
        c1.border = thin_border
        c1.fill = row_fill
        c1.font = Font(name="Segoe UI", size=9)

        c2 = ws.cell(row=row_num, column=2, value=str(cls))
        c2.alignment = center_align
        c2.border = thin_border
        c2.fill = row_fill
        c2.font = Font(name="Segoe UI", size=10, bold=True, color="1F4E78")

        col_cursor = 3
        for day in days:
          for p in periods:
            match = df_result[
                (df_result["الفصل"] == cls)
                & (df_result["اليوم"] == day)
                & (df_result["الحصة"] == p)
            ]
            if not match.empty:
              mat = match.iloc[0]["المادة"]
              tch = match.iloc[0].get("المعلم", "")
              cell_val = f"{mat}\n({tch})" if tch else mat
            else:
              cell_val = "متاحة"

            cell = ws.cell(row=row_num, column=col_cursor, value=cell_val)
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = row_fill
            cell.font = Font(name="Segoe UI", size=8)
            col_cursor += 1

    print(f"📁 تم إنشاء ملف جداول الفصول المجمع بنجاح: {all_classes_matrix_file}")

if __name__ == "__main__":
  generate_timetable()
